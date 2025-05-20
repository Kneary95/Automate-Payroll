import openpyxl
from openpyxl.styles import PatternFill
import os
import glob

# --- Manual alias mapping for known name changes ---
NAME_ALIASES = {
    "Philicia Nichols": "Philicia Haynes",
    "Philicia Haynes": "Philicia Haynes",
    "Beata Blaszczyk": "Reese Blaszczyk",
    "Reese Blaszczyk": "Reese Blaszczyk",
    "Maureen Patterson": "Maureen Schneider",
    "Maureen Schneider": "Maureen Schneider",
    "Tia Claudio": "Female Claudio",
    "Female Claudio": "Tia Claudio",
    # Add more as needed
}

def get_canonical_name(name):
    return NAME_ALIASES.get(name, name)

def normalize_and_clean_gusto(wb):
    if "Gusto Report" not in wb.sheetnames:
        print("Sheet 'Gusto Report' not found.")
        return

    ws = wb["Gusto Report"]
    if "Cleaned Gusto" in wb.sheetnames:
        del wb["Cleaned Gusto"]
    ws_out = wb.create_sheet("Cleaned Gusto")

    headers = ["Clinician Name", "Date", "Job title", "Job hours"]
    ws_out.append(headers)

    row = 1
    max_row = ws.max_row
    out_row = 2

    while row <= max_row:
        cell_value = str(ws.cell(row=row, column=1).value)
        if cell_value.startswith("Hours for "):
            clinician_name = cell_value.replace("Hours for ", "").strip()
            if "," in clinician_name:
                last, first = map(str.strip, clinician_name.split(","))
                clinician_name = f"{first} {last}"

            header_row = [str(ws.cell(row=row + 1, column=col).value).strip() for col in range(1, ws.max_column + 1)]
            j = row + 2
            while j <= max_row:
                entry = str(ws.cell(j, 1).value)
                if entry.startswith("Hours for ") or entry == "":
                    break
                data = {header_row[col - 1]: ws.cell(j, col).value for col in range(1, len(header_row) + 1)}
                base_date = data.get("Date")

                for num in range(1, 25):  
                    title = data.get(f"Job {num} title")
                    hours = data.get(f"Job {num} hours")
                    if title or hours:
                        ws_out.cell(row=out_row, column=1).value = get_canonical_name(clinician_name)
                        ws_out.cell(row=out_row, column=2).value = base_date
                        ws_out.cell(row=out_row, column=2).number_format = 'mm/dd/yyyy'
                        ws_out.cell(row=out_row, column=3).value = title
                        ws_out.cell(row=out_row, column=4).value = hours
                        out_row += 1
                j += 1
            row = j - 1
        row += 1

def get_mapped_service_codes(job_title):
    mappings = {
        "ASSISTANT BEHAVIOR CONSULTANT (ABC -ABA)": ["97153"],
        "BEHAVIOR CONSULTANT": ["H0032", "H2014", "SUB MASTERS", "SUPERVISOR"],
        "BEHAVIOR CONSULTANT - ABA": ["97151", "97155", "97156", "H2014", "97158", "SUB MASTERS", "SUPERVISOR"],
        "BOARD CERTIFIED BEHAVIOR ANALYST": ["97151", "97155", "97156", "97158" "H2014", "SUB MASTERS", "SUPERVISOR"],
        "CLINICAL DIRECTOR": ["CLINICAL DIRECTOR SUPERVISION"],
        "CLINICAL DIRECTOR INDIVIDUAL SERVICES": ["H0032", "H2019"],
        "CLINICAL DIRECTOR SUPERVISION": ["MASTERS SUPERVISION"],
        "MOBILE THERAPY": ["H2019"],
        "MOBILE THERAPY - LICENSED": ["H2019"],
        "TRAINING AND SUPERVISION": [
            "BHT SUPERVISION", "GROUP SUPERVISION", "INDIVIDUAL SUPERVISION", "MASTERS SUPERVISION"
        ],
        "BHT-ABA CENTER": ["97154", "CENTER SUPPORT"],
        "BHT-ABA CENTER GROUP": ["97154", "CENTER SUPPORT"],
        "BEHAVIORAL HEALTH TECHNICIAN": ["H2021", "SUB BHT"],
        "BEHAVIORAL HEALTH TECHNICIAN (BHT-ABA)": ["97153", "SUB BHT"],
    }
    return mappings.get(job_title.strip().upper(), [""])

def compare_cr_to_gusto(wb):
    if "Cleaned Gusto" not in wb.sheetnames or "CR Report" not in wb.sheetnames:
        print("Required sheets not found.")
        return

    ws_gusto = wb["Cleaned Gusto"]
    ws_cr = wb["CR Report"]

    if "Comparison Report" in wb.sheetnames:
        del wb["Comparison Report"]
    ws_out = wb.create_sheet("Comparison Report")

    headers = [
        "Name", "Date", "Job Title (Gusto)", "Service Code (CR)",
        "Gusto Hours", "CR Hours", "Match", "Note"
    ]
    ws_out.append(headers)

    gusto_data = [
        {
            "name": row[0],
            "date": row[1].strftime('%Y-%m-%d') if row[1] else None,
            "job_title": row[2],
            "hours": row[3]
        }
        for row in ws_gusto.iter_rows(min_row=2, values_only=True)
    ]
    cr_data = [
        {
            "name": row[0],
            "service_code": str(row[1]).strip() if row[1] is not None else "",
            "date": row[2].strftime('%Y-%m-%d') if row[2] else None,
            "hours": row[3]
        }
        for row in ws_cr.iter_rows(min_row=2, values_only=True)
    ]

    TOLERANCE = 0.25
    used_cr = [False] * len(cr_data)
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    blue_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    output_row = 2

    for g in gusto_data:
        mapped_codes = get_mapped_service_codes(g.get("job_title", ""))
        mapped_codes_upper = [code.strip().upper() for code in mapped_codes]
        found_match = False
        for i, c in enumerate(cr_data):
            if used_cr[i]:
                continue
            if (
                get_canonical_name(g["name"]) == get_canonical_name(c["name"])
                and g["date"] == c["date"]
                and c["service_code"] is not None
                and c["service_code"].strip().upper() in mapped_codes_upper
            ):
                match = abs((g["hours"] or 0) - (c["hours"] or 0)) <= TOLERANCE
                note = "Matched" if match else "Hour mismatch"
                row_data = [
                    get_canonical_name(g["name"]), g["date"], g["job_title"], c["service_code"],
                    g["hours"], c["hours"], "Yes" if match else "No", note
                ]
                ws_out.append(row_data)
                if not match:
                    for col in range(1, 9):
                        ws_out.cell(row=output_row, column=col).fill = red_fill
                used_cr[i] = True
                found_match = True
                output_row += 1
                break
        if not found_match:
            row_data = [
                get_canonical_name(g["name"]), g["date"], g["job_title"], "", g["hours"], "", "No", "No matching CR entry"
            ]
            ws_out.append(row_data)
            # Highlight blue if job title is Training and Supervision, else red
            if str(g["job_title"]).strip().lower() == "training and supervision":
                for col in range(1, 9):
                    ws_out.cell(row=output_row, column=col).fill = blue_fill
            else:
                for col in range(1, 9):
                    ws_out.cell(row=output_row, column=col).fill = red_fill
            output_row += 1

    for i, c in enumerate(cr_data):
        if not used_cr[i]:
            row_data = [
                get_canonical_name(c["name"]), c["date"], "", c["service_code"], "", c["hours"], "No", "No matching Gusto entry"
            ]
            ws_out.append(row_data)
            for col in range(1, 9):
                ws_out.cell(row=output_row, column=col).fill = red_fill
            output_row += 1

    ws_out.auto_filter.ref = ws_out.dimensions  # Add filter to all columns

    print("✅ Comparison report created.")

def create_discrepancy_summary(wb):
    if "Comparison Report" not in wb.sheetnames:
        print("Comparison Report sheet not found.")
        return

    ws_comp = wb["Comparison Report"]

    # --- Discrepancy Totals (all clinicians, including 0 discrepancies) ---
    if "Discrepancy Totals" in wb.sheetnames:
        del wb["Discrepancy Totals"]
    ws_totals = wb.create_sheet("Discrepancy Totals")
    ws_totals.append(["Clinician Name", "Discrepancy Total"])
    ws_totals.auto_filter.ref = ws_totals.dimensions  # Add filter

    # Build breakdown_dict: {clinician: {note: count}}
    breakdown_dict = {}

    for row in ws_comp.iter_rows(min_row=2, values_only=True):
        name = row[0]
        match = row[6]  # "Match" column
        note = row[7]   # "Note" column
        if name and match == "No":
            if name not in breakdown_dict:
                breakdown_dict[name] = {}
            if note:
                breakdown_dict[name][note] = breakdown_dict[name].get(note, 0) + 1

    # Get all clinicians from the Comparison Report
    all_clinicians = set()
    for row in ws_comp.iter_rows(min_row=2, values_only=True):
        name = row[0]
        if name:
            all_clinicians.add(name)

    for clinician in sorted(all_clinicians):
        total = sum(breakdown_dict.get(clinician, {}).values())
        ws_totals.append([clinician, total])

    ws_totals.auto_filter.ref = ws_totals.dimensions  # Refresh filter after adding data

    # --- Discrepancy Summary (multi-row breakdown) ---
    if "Discrepancy Summary" in wb.sheetnames:
        del wb["Discrepancy Summary"]
    ws_summary = wb.create_sheet("Discrepancy Summary")
    ws_summary.append(["Clinician Name", "Discrepancy Type", "Discrepancy Count"])
    ws_summary.auto_filter.ref = ws_summary.dimensions  # Add filter

    for clinician in sorted(breakdown_dict):
        for note, count in sorted(breakdown_dict[clinician].items()):
            ws_summary.append([clinician, note, count])

    ws_summary.auto_filter.ref = ws_summary.dimensions  # Refresh filter after adding data

    print("✅ Discrepancy summary and totals created.")

# MAIN
if __name__ == "__main__":
    files = glob.glob("*_Payroll.xlsx")
    if not files:
        print("No payroll files found.")
    else:
        latest_file = max(files, key=os.path.getctime)
        print(f"📂 Processing latest payroll file: {latest_file}")
        wb = openpyxl.load_workbook(latest_file)
        normalize_and_clean_gusto(wb)
        compare_cr_to_gusto(wb)
        create_discrepancy_summary(wb)
        for sheet_name in ["CR Report", "Gusto Report", "Cleaned Gusto"]:
            if sheet_name in wb.sheetnames:
                wb[sheet_name].sheet_state = "hidden"
        wb.save(latest_file)
        print(f"✅ All results saved to '{latest_file}'.")
